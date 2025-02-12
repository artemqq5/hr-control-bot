import os
import shutil
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime


class GenerateExcelAnalytics:
    def __init__(self, folder="xlsx_analitics", filename="analitics.xlsx"):
        self.folder = folder
        self.filename = filename
        self.filepath = os.path.join(self.folder, self.filename)

    def _clear_folder(self):
        """Удаляет папку с файлами перед новым сохранением."""
        if os.path.exists(self.folder):
            shutil.rmtree(self.folder)  # Удаление папки и всего содержимого
        os.makedirs(self.folder)  # Создаем заново

    def export(self, data):
        """Генерирует Excel-файл и возвращает его путь."""
        if not data:
            print("Нет даних для експорту.")
            return None

        try:
            # Очищаем папку перед созданием нового файла
            self._clear_folder()

            # Преобразуем данные в DataFrame
            df = pd.DataFrame(data)

            # Переименовываем колонки
            column_mapping = {
                "id": "ID",
                "employee_id": "ID працівника",
                "employee_name": "Працівник",
                "desc": "Опис",
                "user_id": "telegram id керівника",
                "username": "username керівника",
                "created": "Дата створення відомості",
                "realname": "Ім'я керівника",
            }
            df.rename(columns=column_mapping, inplace=True)

            # Форматируем username (добавляем `@`, если не None)
            df["username керівника"] = df["username керівника"].apply(lambda x: f"@{x}" if pd.notna(x) else "—")

            # Форматируем дату
            df["Дата створення відомості"] = df["Дата створення відомості"].apply(
                lambda x: x.strftime("%B %Y, %d %H:%M:%S") if isinstance(x, datetime) else x
            )

            # Открываем книгу
            wb = Workbook()
            ws = wb.active
            ws.title = "Report"

            # === Стили ===
            header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Желтый фон
            header_font = Font(bold=True, color="000000")  # Чёрный жирный текст
            alignment_center = Alignment(horizontal="center", vertical="center")  # Центрирование
            alignment_right = Alignment(horizontal="right", vertical="center")  # Выравнивание чисел вправо
            wrap_text = Alignment(wrap_text=True)  # Перенос строк

            # === Записываем заголовки ===
            for col_num, column_title in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_num, value=column_title)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = alignment_center  # Заголовки центрированы
                ws.column_dimensions[get_column_letter(col_num)].width = max(15, len(column_title) + 2)

            # === Цвета для строк ===
            light_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Белый фон
            dark_fill = PatternFill(start_color="F3F3F3", end_color="F3F3F3", fill_type="solid")  # Светло-серый фон

            # === Записываем данные ===
            for row_num, row_data in enumerate(df.values, 2):
                fill = dark_fill if row_num % 2 == 0 else light_fill
                for col_num, cell_value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                    cell.fill = fill

                    # Выравнивание чисел
                    if isinstance(cell_value, (int, float)):
                        cell.alignment = alignment_right

                    # Центрируем текст в остальных ячейках (если это не число)
                    if isinstance(cell_value, str) and not cell_value.isdigit():
                        cell.alignment = alignment_center

                    # Автоматический перенос длинных текстов (особенно для `Опис`)
                    if isinstance(cell_value, str) and len(cell_value) > 30:
                        cell.alignment = wrap_text
                        ws.column_dimensions[get_column_letter(col_num)].width = 30  # Увеличиваем ширину

            # === Сохраняем файл ===
            wb.save(self.filepath)
            return self.filepath
        except Exception as e:
            print(f"GenerateExcelAnalytics: {e}")
            return None
        